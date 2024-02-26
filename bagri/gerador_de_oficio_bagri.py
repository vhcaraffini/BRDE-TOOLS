from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, PageBreak, PageTemplate, Image
from bagri.funcoes import data_manuscrita, transforma_ponto_em_virgula, configura_paragrafo, cabecalho_e_rodape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus.frames import Frame
from datetime import datetime, timedelta
from reportlab.lib import pagesizes
from reportlab.lib.units import cm
from tkinter import filedialog
from functools import partial
import tkinter as tk
import openpyxl
from windows_functions import get_user_home_folder, create_folder


def gerar_oficio_bagri():
    GET_PATH = get_user_home_folder()
    create_folder(GET_PATH, 'Documentos')
    create_folder(f'{GET_PATH}/Documentos', 'Oficios')
    PATH_IMAGE = 'C:/Users/e.marcus.machado/OneDrive - Banco Regional de Desenvolvimento do Extremo Sul/Imagens'
    
    # Caminho do excel
    root = tk.Tk()
    root.withdraw()

    CAMINHO_ARQUIVO = filedialog.askopenfilename(initialdir="/", title="Selecione um arquivo", filetypes=(("Arquivos do Excel", "*.xlsx"), ("Todos os arquivos", "*.*")))
    workbook = openpyxl.load_workbook(CAMINHO_ARQUIVO)
    sheet_1 = workbook['RESUMO']
    sheet_2 = workbook['REPASSES']

    # PLANILHA 1 (RESUMO)
    # Lista de informações RESUMO
    memorandos = []
    clientes = []
    cnpjs_cpfs = []
    datas_pagamento = []
    datas_vencimento = []
    datas_repasse = []
    data_manuscrita_hoje = data_manuscrita(datetime.now() + timedelta(1))
    convenios_resumo = []

    # Pegando valores do excel e adicionando a lista RESUMO
    for row in sheet_1.iter_rows(values_only=True):
        memorandos.append(row[0])
        datas_pagamento.append(row[1])
        datas_vencimento.append(row[2])
        datas_repasse.append(row[3])
        clientes.append(row[6])
        cnpjs_cpfs.append(row[7])
        convenios_resumo.append(row[4])

    # PLANILHA 2 (REPASSES)
    # Lista de informações REPASSES
    cpfs = []
    mutuarios = []
    repasses = []
    valores_repassados = []
    convenios_repasse = []

    # Pegando valores do excel e adicionando a lista REPASSES
    for row in sheet_2.iter_rows(values_only=True):
        cpfs.append(row[0])
        mutuarios.append(row[1])
        repasses.append(row[2])
        convenios_repasse.append(row[4])
        valores_repassados.append(row[5])

    # Gerando os arquivos baseado naS linhaS (len(clientes))
    for n in range(len(clientes)):
        if n == 0:
            continue
        if clientes[n] == '' or clientes[n] == None:
            break
        else:
            memorando = memorandos[n]
            cliente = clientes[n]
            cnpj_cpf = cnpjs_cpfs[n]
            data_pagamento = datas_pagamento[n].strftime("%d/%m/%Y")
            data_vencimento = datas_vencimento[n].strftime("%d/%m/%Y")
            data_repasse = datas_repasse[n].strftime("%d/%m/%Y")
            convenio = convenios_resumo[n]
        
        if convenio not in convenios_repasse:
            continue

        if convenio == 0 and cliente not in mutuarios:
            continue

        s = 0
        if cliente in mutuarios:
            for mutuario in mutuarios:
                if cliente == mutuario:
                    convenio_avaliador = convenios_repasse[s]
                s += 1

            if convenio_avaliador != 0:
                continue


        estilos = getSampleStyleSheet()

        nome_do_arquivo = F'{GET_PATH}/Documentos/Oficios/Oficio de {cliente}.pdf'

        TAMANHO_DAS_PAGINAS = pagesizes.portrait(pagesizes.A4)

        doc = SimpleDocTemplate(nome_do_arquivo, pagesize=TAMANHO_DAS_PAGINAS, 
            leftMargin = 2.2 * cm, 
            rightMargin = 2.2 * cm,
            topMargin = 3 * cm, 
            bottomMargin = 2 * cm)
        
        frame = Frame(doc.leftMargin, 0, doc.width, doc.height, id='normal')

        conteudo_cabecalho = Image(f'{PATH_IMAGE}/Cabeçalho.jpg', width=21 * cm, height=5 * cm)
        conteudo_rodape = Image(f'{PATH_IMAGE}/Rodapé.jpg', width=21 * cm, height=4 * cm)

        modelo = PageTemplate(id='test', frames=frame, onPage=partial(cabecalho_e_rodape, header_content=conteudo_cabecalho, footer_content=conteudo_rodape))

        doc.addPageTemplates([modelo])

        estilo_justificado = configura_paragrafo(4, 'justificado')
        estilo_a_direita = configura_paragrafo(2, 'direita')
        estilo_centralizado = configura_paragrafo(1, 'centralizado', espacamento=8, espaco_antes=8, espaco_depois=8)
        estilo_assinatura = configura_paragrafo(1, 'assinatura', 9, 6, 6, 6)

        # Adiciona um parágrafo justificado ao PDF
        paragrafos = [
            Paragraph("De:", estilo_justificado),
            Paragraph(f"BRDE/AGCUR/GEARC/SECOB                            Nº {memorando}", estilo_justificado),
            Paragraph("Para:", estilo_justificado),
            Paragraph(f"{cliente}", estilo_justificado),
            Paragraph(f"CNPJ/MF Nº {cnpj_cpf}", estilo_justificado),
            Paragraph(f"", estilo_a_direita),
            Paragraph("Ref.: Reembolso Equalização Programa Banco do Agricultor", estilo_justificado),
            Paragraph(f"", estilo_a_direita),        
            Paragraph("Prezados Senhores:", estilo_justificado),
            Paragraph(f"""Em {data_pagamento} foram pagas por essa Cooperativa parcelas de operações contratadas no âmbito 
            do Programa Banco do Agricultor com vencimento em {data_vencimento}. Após o pagamento, o 
            BRDE enviou o arquivo de solicitação de equalização dessas operações à Fomento Paraná, a 
            qual tendo validado as informações realizou o repasse dos juros a serem equalizados.""", estilo_justificado),
            Paragraph(f"", estilo_a_direita),
            Paragraph(f"No dia {data_repasse} o BRDE efetuou a transferência dos valores constantes na planilha a seguir, diretamente na conta dos mutuários.", estilo_justificado),
            Paragraph(f"", estilo_a_direita),
            Paragraph("Colocamo-nos à disposição para quaisquer esclarecimentos.", estilo_justificado),
            Paragraph(f"", estilo_a_direita),
            Paragraph(f"Curitiba/PR, {data_manuscrita_hoje}.", estilo_a_direita),
            Paragraph(f"", estilo_a_direita),
            Paragraph(f"", estilo_a_direita),
            Paragraph(f"", estilo_a_direita),
            Paragraph(f"____________________________________________", estilo_centralizado),
            Paragraph(f"CRISTIANE MEDIANEIRA DE CASTRO COHEN", estilo_centralizado),
            Paragraph(f"CHEFE DE COBRANÇA E TESOURARIA", estilo_assinatura),
            ]

        # Filtro informações da Tabela
        m = 0
        cpfs_conveniados = []
        mutuario_conveniados = []
        repasse_conveniado = []
        valores_repassados_conveniados = []

        for i in convenios_repasse:
            if i == convenio:
                cpfs_conveniados.append(cpfs[m])
                mutuario_conveniados.append(mutuarios[m])
                repasse_conveniado.append(repasses[m])
                valores_repassados_conveniados.append(valores_repassados[m])
            m += 1

        # Cabeçalho da tabela
        dados_tabela = [['CPF/MF Nº', 'MUTUÁRIO', 'DATA REPASSE', 'VALOR EM R$']]

        # Colunas e linhas da tabela
        if cliente in mutuario_conveniados:
            c = 0
            for p in mutuario_conveniados:
                if cliente == p:
                    linha = [
                        cpfs_conveniados[c], mutuario_conveniados[c], repasse_conveniado[c].strftime("%d/%m/%Y"), 
                        f'R$ {transforma_ponto_em_virgula(valores_repassados_conveniados[c])}'
                        ]
                    dados_tabela.append(linha)
                c += 1

        if cliente not in mutuario_conveniados and convenio != 0:
            for row_count in range(len(mutuario_conveniados)):
                linha = [
                    cpfs_conveniados[row_count], mutuario_conveniados[row_count], repasse_conveniado[row_count].strftime("%d/%m/%Y"), 
                    f'R$ {transforma_ponto_em_virgula(valores_repassados_conveniados[row_count])}'
                    ]
                dados_tabela.append(linha)

        tabela = Table(dados_tabela)

        estilo_tabela = TableStyle([('BACKGROUND', (0, 0), (-1, 0), 'grey'),
                                    ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
                                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                    ('BACKGROUND', (0, 1), (-1, -1), 'white'),
                                    ('GRID', (0, 0), (-1, -1), 1, 'black')])
        
        tabela.setStyle(estilo_tabela)
        tabela.setStyle(estilo_tabela)

        # Adicionando a proximas paginas
        paragrafos.append(PageBreak())
        paragrafos.append(tabela)

        doc.build(paragrafos)